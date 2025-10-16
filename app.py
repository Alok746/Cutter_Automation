from flask import Flask, session, redirect, url_for, request, render_template, jsonify, send_file, Response, stream_with_context
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import load_workbook
from dotenv import load_dotenv
from openai import OpenAI

import datetime
import json
import os
import pandas as pd
import re
from copy import deepcopy

# Import processors for question types
from routes.single_choice import process_single_choice, get_single_choice_data, process_single_choice_qualtrics
from routes.matrix_question import process_matrix_question, get_matrix_question_data, process_matrix_question_qualtrics
from routes.multiple_select import process_multi_select, get_multi_select_data, process_multi_select_qualtrics
from routes.cross_cut import process_cross_cut, get_cross_cut_data, process_cross_cut_qualtrics
from routes.rank_based import process_ranked_question ,get_ranked_data, process_ranked_question_qualtrics
from routes.nps_question import process_nps_question, get_nps_data, process_nps_question_qualtrics
from routes.share_of_wallet import process_share_of_wallet, get_sow_data
from routes.ppt_export import build_matrix_slide,build_nps_slide,build_single_choice_slide
from routes.rag_chatbot import SurveyRAG
from routes.sql_agent_utils import run_sql_agent

from utils import detect_nps_questions, detect_single_choice_questions,detect_multi_select_questions, apply_global_filters, detect_matrix_questions_qualtrics, detect_multi_select_questions_qualtrics,detect_nps_questions_qualtrics,detect_single_choice_questions_qualtrics, extract_answer_values

load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

print("🔑 OpenAI key loaded:", os.getenv("OPENAI_API_KEY")[:10] + "...")

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.secret_key = "bleh"

# -----------------------------------------
# Combined upload + sheet listing
# -----------------------------------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'GET':
        # 🧹 Clear any existing session data
        session.clear()       
    if request.method == 'POST':
        file = request.files.get('excel_file')
        if file:
            filename = file.filename
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            xls = pd.ExcelFile(filepath)
            sheet_names = [s.strip() for s in xls.sheet_names]

            # Detect if Qualtrics-style file is present
            is_qualtrics = "variable information" in [s.lower() for s in sheet_names]

            # Exclude Answer key and Variable info from dropdown
            excluded = {"answer key", "variable information"}
            sheets = [s for s in sheet_names if s.strip().lower() not in excluded]

            return jsonify({
                "success": True,
                "filename": filename,
                "sheets": sheets,
                "is_qualtrics": is_qualtrics
            })
        return jsonify({"success": False, "error": "No file uploaded"})
    return render_template('index.html')

# -----------------------------------------
# Handles mode (single vs multi) after upload+sheet
# -----------------------------------------
@app.route('/route_selector', methods=['POST'])
def route_selector():
    action = request.form.get("action")

    if action == "survey":
        filename = request.form['filename']
        sheet = request.form['sheet']
        data_format = request.form.get('data_format', 'inquery')  # Default to inquery
        return select_columns(filename, sheet, data_format)
    elif action == "genai":
        return index()

    return "Invalid action", 400


# -----------------------------------------
# Shows question list after selecting sheet
# -----------------------------------------
@app.route('/select_columns', methods=['GET','POST'])
def select_columns(filename=None, sheet=None, data_format=None):
    if request.method == 'GET':
        filename = session.get('filename')
        sheet = session.get('sheet')

        if not data_format or data_format.strip() == "":
            data_format = request.form.get('data_format', 'inquery')

        if not filename or not sheet:
            return redirect(url_for('index'))
    else:
        filename = request.form['filename']
        sheet = request.form['sheet']
        session['filename'] = filename
        session['sheet'] = sheet
        data_format = request.form.get('data_format', 'inquery')

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    df_full = pd.read_excel(filepath, sheet_name=sheet, header=2)
    df = df_full.iloc[1:].reset_index(drop=True) if data_format == "qualtrics" else df_full
    df_key = pd.read_excel(filepath, sheet_name="Answer key", header=None)

    # ----------------------------------------------------------------------------------
    # Qualtrics-specific question text extraction (Variable Information sheet + fallback)
    # ----------------------------------------------------------------------------------
    question_map = {}
    if data_format == "qualtrics":
        variable_text_map = {}
        try:
            df_varinfo = pd.read_excel(filepath, sheet_name="Variable information", header=None)
            from collections import defaultdict
            grouped = defaultdict(list)

            for _, row in df_varinfo.iterrows():
                if pd.notna(row[0]) and pd.notna(row[2]):
                    raw = str(row[0]).strip()
                    txt = str(row[2]).strip()
                    m = re.match(r"^(Q\d+)_\d+$", raw)
                    if m:
                        grouped[m.group(1)].append(txt)

            def lcp(strs):
                if not strs: return ""
                s1, s2 = min(strs), max(strs)
                for i, ch in enumerate(s1):
                    if i >= len(s2) or ch != s2[i]:
                        return s1[:i].rstrip(" :-–")
                return s1.rstrip(" :-–")

            for qid, texts in grouped.items():
                variable_text_map[qid] = lcp(texts)

        except Exception as e:
            print("Error reading Variable information:", e)

        question_map.update(variable_text_map)

        # Fallback to header row for missing QIDs
        text_row = df_full.iloc[0]
        for col in df_full.columns:
            if not isinstance(col, str):
                continue
            m = re.match(r"^(Q\d+)", col)
            if not m:
                continue
            qid = m.group(1)
            if qid in question_map:
                continue

            pattern = re.compile(rf"^{re.escape(qid)}(?:[_:]\d+)?$")
            cols = [c for c in df_full.columns if pattern.match(c)]
            texts = [str(text_row[c]).strip() for c in cols]

            if len(texts) == 1:
                question_map[qid] = texts[0]
            else:
                prefixes = [t.split(" - ")[0] for t in texts]
                question_map[qid] = prefixes[0] if len(set(prefixes)) == 1 else lcp(texts)

    # ----------------------------------------------------------------------------------
    # InQuery-style question mapping (keep this from your original file)
    # ----------------------------------------------------------------------------------
    else:
        data_columns = set()
        for col in df.columns:
            if isinstance(col, str):
                match = re.search(r"(Q\d+)", col)
                if match:
                    data_columns.add(match.group(1))

        i = 0
        while i < len(df_key):
            row = df_key.iloc[i]
            if pd.notna(row[0]) and pd.notna(row[1]):
                qid_match = re.match(r"(Q\d+)", str(row[0]).strip())
                if qid_match:
                    qid = qid_match.group(1)
                    text = str(row[1]).strip()

                    has_options = False
                    j = i + 1
                    while j < len(df_key):
                        subrow = df_key.iloc[j]
                        if pd.isna(subrow[0]):
                            break
                        if pd.notna(subrow[1]):
                            has_options = True
                            break
                        j += 1

                    if has_options:
                        question_map[qid] = text
            i += 1

        # Only include questions that exist in the data columns
        question_map = {qid: text for qid, text in question_map.items() if qid in data_columns}

    # ----------------------------------------------------------------------------------
    # Final sorting and recommendation tagging
    # ----------------------------------------------------------------------------------
    columns = sorted(question_map.keys(), key=lambda q: int(q[1:]))
    question_pairs = [(qid, question_map[qid]) for qid in columns]

    if data_format == 'qualtrics':
        nps_recs = detect_nps_questions_qualtrics(df)
        sc_recs = detect_single_choice_questions_qualtrics(df)
        ms_recs = detect_multi_select_questions_qualtrics(df)
        mx_recs = detect_matrix_questions_qualtrics(df)
        recommendations = {qid: [] for qid, _ in question_pairs}
        for q in nps_recs:
            if q in recommendations: recommendations[q].append("nps")
        for q in sc_recs:
            if q in recommendations: recommendations[q].append("single_choice")
        for q in ms_recs:
            if q in recommendations: recommendations[q].append("multi_select")
        for q in mx_recs:
            if q in recommendations: recommendations[q].append("matrix")
    else:
        nps_recs = detect_nps_questions(df, df_key)
        sc_recs = detect_single_choice_questions(df, df_key)
        ms_recs = detect_multi_select_questions(df)
        recommendations = {qid: [] for qid, _ in question_pairs}
        for q in nps_recs:
            if q in recommendations: recommendations[q].append("nps")
        for q in sc_recs:
            if q in recommendations: recommendations[q].append("single_choice")
        for q in ms_recs:
            if q in recommendations: recommendations[q].append("multi_select")

    return render_template(
        'select_columns.html',
        filename=filename,
        sheet=sheet,
        question_pairs=question_pairs,
        recommendations=recommendations,
        active_tab='survey',
        data_format=data_format
    )

@app.route('/genai')
def genai():
    filename = session.get("filename")
    sheet = session.get("sheet")

    if not filename or not sheet:
        # No data uploaded → redirect user back to upload
        return redirect(url_for('index'))

    # Optional: preload embeddings once (can skip for faster load)
    try:
        rag = SurveyRAG(app.config['UPLOAD_FOLDER'], filename, sheet)
        rag.load_excel()
        rag.convert_to_documents(max_rows=None)
        print("🧪 Rows passed to document encoder:", len(rag.documents))
        print("🧪 Debug: total rows in RAG dataframe:", len(rag.df))
        rag.embed_documents()
        rag.build_faiss_index()
    except Exception as e:
        return f"❌ Error preparing AI assistant: {str(e)}", 500

    return render_template(
        'excel_genai.html',
        filename=filename,
        sheet=sheet,
        active_tab='genai'
    )
       
@app.route('/chat', methods=['POST'])
def chat():
    user_query = request.get_json().get("message")
    filename = session.get("filename")
    sheet = session.get("sheet")

    if not all([user_query, filename, sheet]):
        return jsonify({'error': 'Missing input or session data'}), 400

    def generate():
        try:
            # Load and prepare RAG context
            rag = SurveyRAG(app.config['UPLOAD_FOLDER'], filename, sheet)
            rag.load_excel()
            rag.convert_to_documents(max_rows=None)
            
            prompt = """
                You are a helpful assistant analyzing structured survey data. Answer each user question clearly and concisely, using only the provided survey context.

                📌 Format your response in **Markdown** with proper structure:
                - Use `###` for section headings
                - Use one bullet (`-`) per item
                - Start each point on a new bullet line. Never include multiple bold labels in the same bullet.
                - No extra line breaks between bullets
                - Do not insert `**` or `:` on separate lines
                - Do not use numbering like `1.`, `2.`, `3.`
                - ✅ If the user asks for a **number**, **count**, or **specific metric**, provide the most relevant numeric value available from the context.
                - ✅ Include exact values (e.g., counts, percentages) if they're directly relevant to the question.

                ✅ Example:

                ### Key Strengths of ABB
                - **Product Quality:** ABB is known for good quality.
                - **Service:** Respondents praise their follow-up.
                - **Pricing:** Considered affordable.

                For **verbatim** or open-ended questions:
                - Identify columns in Raw data that contain **free-text answers** (not coded).
                - Read and analyze all verbatim entries to answer the user query appropriately (e.g., summarize sentiment, highlight key mentions, extract themes).

                ⚠️ Always match options using the answer key before interpreting raw numeric values.
                ⚠️ Do not guess codes. Only count if there is a confirmed mapping in the Answer Key.
                """         
            system_prompt = prompt
            
            raw_df = rag.raw_df
            key_df = rag.key_df

            # 🔍 Very simple SQL-style query detection
            user_lower = user_query.lower()

            # 🧮 SQL-style detection
            is_sql_like = any(
                kw in user_lower
                for kw in ["how many", "count", "percentage", "most", "top", "group by", "average", "filter", "sum", "respondents who"]
            )

            # 💬 Verbatim-style detection
            is_verbatim = any(
                kw in user_lower
                for kw in ["what do people say", "good qualities", "feedback", "comments", "open-ended", "suggestions", "improvements"]
            )

            if is_sql_like:
                # 🧠 Use SQL Agent flow
                sql_response = run_sql_agent(user_query, raw_df, key_df)

                # Stream result as a text block (you can later upgrade to table if it's a DataFrame)
                for line in str(sql_response).splitlines():
                    yield f"data: {line}\n\n"
                return
            
            if is_verbatim:
                # 🔁 Fall back to RAG for verbatim-style answers
                rag.embed_documents()
                rag.build_faiss_index()

                top_chunks = rag.retrieve(user_query, top_k=5)
                context = "\n---\n".join(top_chunks)

                messages = [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"Survey Context:\n{context}\n\nQuestion: {user_query}"}
                ]

                with client.chat.completions.create(
                    model="gpt-4o",
                    messages=messages,
                    temperature=0.1,
                    stream=True,
                ) as stream:
                    for chunk in stream:
                        content = chunk.choices[0].delta.content or ""
                        yield f"data: {content}\n\n"
                return

            # 🧠 Else fallback to existing RAG-based logic
            rag.embed_documents()
            rag.build_faiss_index()

            top_chunks = rag.retrieve(user_query, top_k=5)
            context = "\n---\n".join(top_chunks)

            

            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Survey Context:\n{context}\n\nQuestion: {user_query}"}
            ]

            with client.chat.completions.create(
                model="gpt-4o",
                messages=messages,
                temperature=0.1,
                stream=True,
            ) as stream:
                for chunk in stream:
                    content = chunk.choices[0].delta.content or ""
                    yield f"data: {content}\n\n"

        except Exception as e:
            import traceback
            traceback.print_exc()
            yield f"data: [ERROR] {str(e)}\n\n"

    return Response(stream_with_context(generate()), content_type="text/event-stream")
# -----------------------------------------
# Renders selected question results
# -----------------------------------------
def get_question_text_from_key(df_key, qid):
    for _, row in df_key.iterrows():
        if pd.notna(row[0]) and str(row[0]).strip() == qid:
            return str(row[1]).strip() if pd.notna(row[1]) else qid
    return qid

def get_question_text_from_raw(df_full: pd.DataFrame, qid: str) -> str:
    """
    Return the human-readable text for question `qid` from df_full,
    which must have been loaded with header=row_of_Qcodes (so df_full.columns
    are like 'Q1', 'Q2_1', 'Q2_2', …) and df_full.iloc[0] holds the question texts.
    """
    # 1) The row immediately below the header holds the question text
    text_row = df_full.iloc[0]

    # 2) Find all columns that start with this qid (e.g. 'Q2', 'Q2_1', 'Q2:sub', etc.)
    pattern = re.compile(rf'^{re.escape(qid)}(?:[_:]\d+)?$')
    cols = [c for c in df_full.columns if isinstance(c, str) and pattern.match(c)]
    if not cols:
        return qid  # no match → give back the qid itself

    # 3) Pull their texts
    texts = [ str(text_row[c]).strip() for c in cols ]

    # 4) If only one column, that's the text
    if len(texts) == 1:
        return texts[0]

    # 5) If multiple (e.g. multi-select), return the common prefix before ' - '
    prefixes = [t.split(' - ')[0] for t in texts]
    if len(set(prefixes)) == 1:
        return prefixes[0]

    # 6) Otherwise fall back to the longest common prefix of all full texts
    shortest, longest = min(texts), max(texts)
    for i, ch in enumerate(shortest):
        if ch != longest[i]:
            return shortest[:i].rstrip()
    return shortest.rstrip()

@app.route('/compare_multi_questions', methods=['POST'])
def compare_multi_questions():
    filename = request.form['filename']
    sheet = request.form['sheet']
    data_format = request.form.get('data_format', 'inquery')
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    df_full = pd.read_excel(filepath, sheet_name=sheet, header=2)
    df = df_full.iloc[1:].reset_index(drop=True) if data_format == "qualtrics" else df_full
    df_key = pd.read_excel(filepath, sheet_name="Answer key", header=None)

    # ------------------------
    # Filter inputs
    # ------------------------
    filter_count = int(request.form.get('filter_count', 0))
    filter_questions = []
    filter_values = []
    for i in range(filter_count):
        fq = request.form.get(f'filter_question_{i}')
        fv = request.form.get(f'filter_value_{i}', '__all__')
        if fq:
            filter_questions.append(fq)
            filter_values.append(fv)

    base_filters = {
        'filter_questions': filter_questions,
        'filter_values': filter_values,
        'sort_column': request.form.get('sort_column', ''),
        'data_format': data_format
    }

    questions = [
        key.replace("include_", "")
        for key in request.form.keys()
        if key.startswith("include_") and request.form.get(key)
    ]

    results = []
    all_columns = []

    for q in questions:
        types = list(dict.fromkeys(request.form.getlist(f"type_{q}")))
        for q_type in types:
            filters = deepcopy(base_filters)
            filters['cut_column'] = request.form.get(f'cut_column_{q}', '')
            filters['max_rank'] = request.form.get(f'max_rank_{q}', '')

            try:
                # ----------- Core branching by type and data format ------------
                if q_type == 'cross_cut' and filters['cut_column']:
                    if data_format == "qualtrics":
                        html = process_cross_cut_qualtrics(filepath, sheet, q, filters)
                    else:
                        html = process_cross_cut(filepath, sheet, q, filters)
                    question_text = f"Cross cut {q} x {filters['cut_column']} - Cross Cut Summary"

                elif q_type == 'single_choice':
                    if data_format == "qualtrics":
                        html = process_single_choice_qualtrics(filepath, sheet, q, filters)
                        question_text = get_question_text_from_raw(df_full, q)
                    else:
                        html = process_single_choice(filepath, sheet, q, filters)
                        question_text = get_question_text_from_key(df_key, q)

                elif q_type == 'matrix':
                    if data_format == "qualtrics":
                        html = process_matrix_question_qualtrics(filepath, sheet, q, filters)
                        question_text = get_question_text_from_raw(df_full, q)
                    else:
                        html = process_matrix_question(filepath, sheet, q, filters)
                        question_text = get_question_text_from_key(df_key, q)

                elif q_type == 'multi_select':
                    if data_format == "qualtrics":
                        html = process_multi_select_qualtrics(filepath, sheet, q, filters)
                        question_text = get_question_text_from_raw(df_full, q)
                    else:
                        html = process_multi_select(filepath, sheet, q, filters)
                        question_text = get_question_text_from_key(df_key, q)

                elif q_type == 'ranked':
                    if data_format == "qualtrics":
                        html = process_ranked_question_qualtrics(filepath, sheet, q, filters)
                        question_text = get_question_text_from_raw(df_full, q)
                    else:
                        html = process_ranked_question(filepath, sheet, q, filters)
                        question_text = get_question_text_from_key(df_key, q)

                elif q_type == 'nps':
                    if data_format == "qualtrics":
                        html = process_nps_question_qualtrics(filepath, sheet, q, filters)
                        question_text = get_question_text_from_raw(df_full, q)
                    else:
                        html = process_nps_question(filepath, sheet, q, filters)
                        question_text = get_question_text_from_key(df_key, q)

                elif q_type == 'sow':
                    html = process_share_of_wallet(filepath, sheet, q, filters)
                    question_text = get_question_text_from_key(df_key, q)

                else:
                    html = f"<p>Unsupported or incomplete type for {q}</p>"
                    question_text = f"{q} - {q_type}"

                results.append({'question': q, 'text': question_text, 'html': html})
                all_columns.append(q)

            except Exception as e:
                results.append({
                    'question': q,
                    'text': f"{q} - {q_type}",
                    'html': f"<p>Error processing {q} ({q_type}): {str(e)}</p>"
                })

    results.sort(key=lambda r: int(r['question'][1:]) if r['question'].startswith('Q') else 999)

    # -----------------------------
    # Build filterable column list
    # -----------------------------
    filter_columns = []
    seen = set()
    for _, row in df_key.iterrows():
        if pd.notna(row[0]) and re.match(r'^Q\d+$', str(row[0]).strip()):
            code = str(row[0]).strip()
            label = str(row[1]).strip() if pd.notna(row[1]) else code
            if code not in seen:
                filter_columns.append((code, label))
                seen.add(code)

    # -----------------------------
    # Recommendation tagging
    # -----------------------------
    recommendations = {code: [] for code, _ in filter_columns}
    if data_format == 'qualtrics':
        nps_recs = detect_nps_questions_qualtrics(df)
        sc_recs = detect_single_choice_questions_qualtrics(df)
        ms_recs = detect_multi_select_questions_qualtrics(df)
        mx_recs = detect_matrix_questions_qualtrics(df)
        for q in nps_recs:
            if q in recommendations: recommendations[q].append("nps")
        for q in sc_recs:
            if q in recommendations: recommendations[q].append("single_choice")
        for q in ms_recs:
            if q in recommendations: recommendations[q].append("multi_select")
        for q in mx_recs:
            if q in recommendations: recommendations[q].append("matrix")
    else:
        nps_recs = detect_nps_questions(df, df_key)
        sc_recs = detect_single_choice_questions(df, df_key)
        ms_recs = detect_multi_select_questions(df)
        for q in nps_recs:
            if q in recommendations: recommendations[q].append("nps")
        for q in sc_recs:
            if q in recommendations: recommendations[q].append("single_choice")
        for q in ms_recs:
            if q in recommendations: recommendations[q].append("multi_select")

    # -----------------------------
    # If Qualtrics: override filter labels using VarInfo
    # -----------------------------
    question_text_map = {}
    if data_format == 'qualtrics':
        try:
            df_varinfo = pd.read_excel(filepath, sheet_name="Variable information", header=None)
            from collections import defaultdict
            grouped = defaultdict(list)

            for _, row in df_varinfo.iterrows():
                if pd.notna(row[0]) and pd.notna(row[2]):
                    qid = str(row[0]).strip()
                    if re.match(r"^Q\d+(_\d+)?$", qid):
                        base_qid = qid.split("_")[0]
                        grouped[base_qid].append(str(row[2]).strip())

            def lcp(strs):
                if not strs: return ""
                s1, s2 = min(strs), max(strs)
                for i, ch in enumerate(s1):
                    if i >= len(s2) or ch != s2[i]:
                        return s1[:i].rstrip(" :-–")
                return s1.rstrip(" :-–")

            for base_qid, texts in grouped.items():
                question_text_map[base_qid] = lcp(texts)

            filter_columns = [
                (code, question_text_map.get(code, fallback))
                for code, fallback in filter_columns
            ]
        except Exception as e:
            print("Error reading VarInfo for filter labels:", e)

    return render_template(
        'display_multi_results.html',
        results=results,
        filename=filename,
        sheet=sheet,
        all_columns=all_columns,
        filter_columns=filter_columns,
        filter_questions=filter_questions,
        filter_values=filter_values,
        sort_column=base_filters['sort_column'],
        data_format=data_format,
        recommendations=recommendations,
        question_text_map=question_text_map
    )

# -----------------------------------------
# AJAX endpoint to load filter values
# -----------------------------------------
@app.route("/get_answer_key_values", methods=["POST"])
def get_answer_key_values():
    filename = request.form.get("filename")
    question = request.form.get("question")
    data_format = request.form.get("data_format", "inquery")
    
    if not filename or not question:
        return jsonify([])

    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        df_key = pd.read_excel(filepath, sheet_name="Answer key", header=None)

        values = extract_answer_values(df_key, question, data_format)
        return jsonify(values)

    except Exception as e:
        print("Error in get_answer_key_values:", e)
        return jsonify([])

# -----------------------------------------
# AJAX endpoint to download excel
# -----------------------------------------

@app.route('/download_all_excel', methods=['POST'])
def download_all_excel():
    filename = request.form.get('filename')
    sheet = request.form.get('sheet')
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    question_ids = request.form.getlist("questions")

    buffer = BytesIO()
    writer = pd.ExcelWriter(buffer, engine='openpyxl')
    question_text_map = {}

    for qid in question_ids:
        types = request.form.getlist(f"type_{qid}")
        filters = {
            'filter_questions': request.form.getlist('filter_question_0'),
            'filter_values': request.form.getlist('filter_value_0'),
            'sort_column': request.form.get('sort_column', '')
        }
        filters['cut_column'] = request.form.get(f'cut_column_{qid}', '')
        filters['max_rank'] = request.form.get(f'max_rank_{qid}', '')

        for qtype in types:
            try:
                # Load appropriate data
                if qtype == 'single_choice':
                    output = get_single_choice_data(filepath, sheet, qid, filters)
                    df1 = pd.DataFrame(output["response_summary"], columns=["Option", "Count", "Percentage"])
                    df2 = None

                elif qtype == 'multi_select':
                    output = get_multi_select_data(filepath, sheet, qid, filters)
                    df1 = pd.DataFrame(output["response_summary"],
                                       columns=["Option", "Count", "% Respondents", "% Responses"])
                    df2 = None

                elif qtype == 'matrix':
                    output = get_matrix_question_data(filepath, sheet, qid, filters)
                    df1 = pd.DataFrame(output["count_matrix"], columns=output["col_labels"])
                    df1.insert(0, "Label", output["row_labels"])
                    df2 = pd.DataFrame(output["percent_matrix"], columns=output["col_labels"])
                    df2.insert(0, "Label", output["row_labels"])

                elif qtype == 'ranked':
                    output = get_ranked_data(filepath, sheet, qid, filters)
                    df1 = pd.DataFrame([[label, total] + counts for label, total, counts in output["result_matrix"]],
                                       columns=["Label", "Total"] + output["rank_labels"])
                    df2 = pd.DataFrame([[label, overall] + row for label, overall, row in output["percent_matrix"]],
                                       columns=["Label", "Overall %"] + output["rank_labels"])

                elif qtype == 'cross_cut':
                    output = get_cross_cut_data(filepath, sheet, qid, filters)
                    df1 = pd.DataFrame([[label, total] + row for label, total, row in output["result_matrix"]],
                                       columns=["Label", "Total"] + output["cut_headers"])
                    df2 = pd.DataFrame([[label, overall] + row for label, overall, row in output["percent_matrix"]],
                                       columns=["Label", "Overall %"] + output["cut_headers"])

                elif qtype == 'nps':
                    output = get_nps_data(filepath, sheet, qid, filters)
                    df1 = pd.DataFrame([[score] + row for score, row in output["result_matrix"]],
                                       columns=["Score"] + output["col_labels"])
                    df2 = pd.DataFrame([
                        ["Promoters"] + output["promoters"],
                        ["Neutrals"] + output["neutrals"],
                        ["Detractors"] + output["detractors"],
                        ["NPS Score (%)"] + [val / 100 for val in output["nps_scores"]],
                        ["Average Score"] + output["average_scores"]
                    ], columns=["Category"] + output["col_labels"])

                elif qtype == 'sow':
                    output = get_sow_data(filepath, sheet, qid, filters)
                    df1 = pd.DataFrame(output["count_matrix"],
                                       index=output["row_labels"],
                                       columns=output["col_labels"]).reset_index().rename(columns={"index": "Range"})
                    df2 = None

                else:
                    df1 = pd.DataFrame([["Unsupported question type"]])
                    df2 = None

                sheet_name = f"{qid}_{qtype}"[:31]
                question_text = output.get("question_text", f"{qid}: {qtype}")
                question_text_map[sheet_name] = question_text

                # Write count table
                df1_start = 2
                df1.to_excel(writer, sheet_name=sheet_name, index=False, startrow=df1_start)

                # Write percentage table
                df2_start = df1_start + len(df1) + 3 if df2 is not None else None
                if df2 is not None:
                    df2.to_excel(writer, sheet_name=sheet_name, index=False, startrow=df2_start)

            except Exception as e:
                sheet_name = f"{qid}_{qtype}"[:31]
                question_text_map[sheet_name] = f"Error processing {qid}: {str(e)}"
                pd.DataFrame([[f"Error processing {qid}: {str(e)}"]]).to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()
    buffer.seek(0)

    wb = load_workbook(buffer)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_rule = ColorScaleRule(start_type='min', start_color='e5f4e2',
                                mid_type='percentile', mid_value=50, mid_color='8fce93',
                                end_type='max', end_color='1e7b33')

    def apply_formatting(ws, df_start, df_len, has_percentage=False):
        # Bold header row
        for cell in ws[df_start]:
            cell.font = Font(bold=True)

        # Borders + formatting
        for row in ws.iter_rows(min_row=df_start+1, max_row=df_start+df_len):
            for idx, cell in enumerate(row, start=1):
                cell.border = border
                # Apply % format where appropriate
                if has_percentage and isinstance(cell.value, float) and 0 <= cell.value <= 1 and idx > 1:
                    cell.number_format = '0.0%'

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        question_text = question_text_map.get(sheet, sheet)
        ws.cell(row=1, column=1).value = question_text
        ws.cell(row=1, column=1).font = Font(bold=True, italic=True)

        # Read row lengths
        df1_start = 2
        df1_len = 0
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Label" or ws.cell(row=r, column=1).value == "Range":
                break
            df1_len += 1
        df2_start = df1_start + df1_len + 3
        df2_len = ws.max_row - df2_start if ws.max_row > df2_start else 0

        apply_formatting(ws, df1_start, df1_len)
        if df2_len > 1:
            apply_formatting(ws, df2_start, df2_len, has_percentage=True)
            for col_idx in range(2, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                ws.conditional_formatting.add(
                    f"{col_letter}{df2_start+1}:{col_letter}{df2_start+df2_len}",
                    green_rule
                )

        # Auto column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        final_buffer,
        as_attachment=True,
        download_name=f"survey_results_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ------------------------------------------------------
# AJAX endpoint to download thinkcell enables Powerpoint
# ------------------------------------------------------

@app.route('/download_thinkcell_ppt', methods=['POST'])
def download_thinkcell_ppt():
    filename = request.form.get("filename")
    sheet = request.form.get("sheet")
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)

    filters = {
        "filter_questions": [],
        "filter_values": [],
        "cut_column": "",
        "max_rank": "",
        "sort_column": request.form.get("sort_column", "")
    }

    # Define which template to use for each chart type
    template_map = {
        "nps": "NPS.pptx",
        "matrix": "Matrix.pptx",
        "single_choice": "Single.pptx"
    }

    questions = request.form.getlist("questions")
    blocks_by_type = {
        "nps": [],
        "matrix": [],
        "single_choice": []
    }

    single_choice_qids = []

    for qid in questions:
        types = request.form.getlist(f"type_{qid}")
        filters["cut_column"] = request.form.get(f'cut_column_{qid}', '')
        filters["max_rank"] = request.form.get(f'max_rank_{qid}', '')

        try:
            if "single_choice" in types:
                single_choice_qids.append(qid)
                continue

            elif "nps" in types:
                chart_data = build_nps_slide(filepath, sheet, qid, filters)
                blocks_by_type["nps"].append({
                    "slide": 1,
                    "data": [{"name": k, "table": v} for k, v in chart_data.items()]
                })

            elif "matrix" in types:
                chart_data = build_matrix_slide(filepath, sheet, qid, filters)
                blocks_by_type["matrix"].append({
                    "slide": 1,
                    "data": [{"name": k, "table": v} for k, v in chart_data.items()]
                })

        except Exception as e:
            print(f"❌ Failed to build {types[0]} for {qid}: {e}")

    # ✅ Combine single choice questions into one chart
    from collections import defaultdict

    # 🔁 Group single choice QIDs based on submitted group numbers
    grouped_qids = defaultdict(list)
    for qid in single_choice_qids:
        group_val = request.form.get(f"group_single_{qid}", "1")  # default group = 1
        grouped_qids[group_val].append(qid)

    # 🔄 Generate separate charts per group
    for group_id, qid_group in grouped_qids.items():
        try:
            chart_data = build_single_choice_slide(filepath, sheet, qid_group, filters)
            blocks_by_type["single_choice"].append({
                "slide": 1,
                "data": [{"name": k, "table": v} for k, v in chart_data.items()]
            })
        except Exception as e:
            print(f"❌ Failed to build single_choice group {group_id}: {e}")

    # ✅ Now assign templates to first block of each chart type
    slide_blocks = []

    for chart_type, blocks in blocks_by_type.items():
        if not blocks:
            continue

        for block in blocks:
            block["template"] = template_map[chart_type]
            slide_blocks.append(block)

    if not slide_blocks:
        return "⚠️ No valid charts to export.", 400

    # ✅ Export as ThinkCell-compatible .ppttc
    from io import BytesIO
    buffer = BytesIO()
    buffer.write(json.dumps(slide_blocks, indent=2).encode("utf-8"))
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="survey_output.ppttc",
        mimetype="application/json"
    )

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True)